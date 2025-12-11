import os
import json
import re
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

load_dotenv()

from langchain_ollama import OllamaLLM, OllamaEmbeddings
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from langchain_postgres import PGVector
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import TYPE_STRING

from models import Rule, Datatype, ExtractionResponse, GenerationRequest, IntermediateVariable, HelperRuleDefinition
from prompts import (
    ENRICHMENT_PROMPT_TEMPLATE, 
    DATATYPE_GENERATION_PROMPT_TEMPLATE,
    SPREADSHEET_GENERATION_PROMPT_TEMPLATE,
    DECISION_TABLE_GENERATION_PROMPT_TEMPLATE,
    TEST_GENERATION_PROMPT_TEMPLATE,
    ORCHESTRATOR_PROMPT_TEMPLATE
)

# Configuration
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
LLM_MODEL = os.getenv("LLM_MODEL", "gpt-oss:20b")
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "mxbai-embed-large:latest")

# Database Config
DB_USER = os.getenv("PG_USER", "user")
DB_PASSWORD = os.getenv("PG_PASSWORD", "password")
DB_HOST = os.getenv("PG_HOST", "localhost")
DB_PORT = os.getenv("PG_PORT", "5432")
DB_NAME = os.getenv("PG_DB", "openl_rag")
CONNECTION_STRING = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

class GenerationService:
    def __init__(self):
        self.llm = OllamaLLM(base_url=OLLAMA_BASE_URL, model=LLM_MODEL)
        self.embeddings = OllamaEmbeddings(base_url=OLLAMA_BASE_URL, model=EMBEDDING_MODEL)
        self.vector_store = PGVector(
            embeddings=self.embeddings,
            collection_name="openl_guide",
            connection=CONNECTION_STRING,
            use_jsonb=True,
        )

    def _get_rag_context(self, query: str) -> str:
        """Retrieve relevant context from the vector store."""
        try:
            retriever = self.vector_store.as_retriever(
                search_type="mmr",
                search_kwargs={"k": 5, "fetch_k": 10}
            )
            docs = retriever.invoke(query)
            return "\n\n".join([doc.page_content for doc in docs])
        except Exception as e:
            print(f"RAG Retrieval failed: {e}")
            return ""

    async def enrich_rules(self, rules: List[Any], text: str, filename: Optional[str] = None) -> ExtractionResponse:
        """
        Architect Phase: Analyze rules and define the 3-layer structure.
        Uses a local cache (enrich_cache/) if logic has been generated for this file before.
        """
        # 0. Cache Lookup
        CACHE_DIR = "enrich_cache"
        existing_context_str = ""
        cached_data = None
        cache_path = None
        
        if filename:
            os.makedirs(CACHE_DIR, exist_ok=True)
            # Sanitize filename (basic)
            safe_name = re.sub(r'[^a-zA-Z0-9_\-\.]', '_', filename)
            cache_path = os.path.join(CACHE_DIR, f"{safe_name}.json")
            
            if os.path.exists(cache_path):
                print(f"[CACHE] Hit for {filename}")
                try:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        cached_data = json.load(f)
                    
                    # Format for Prompt
                    existing_datatypes = cached_data.get("datatypes", [])
                    existing_vars = cached_data.get("intermediate_variables", [])
                    
                    dt_str = "\n".join([f"- Datatype {d['name']}: {[f['name'] for f in d['fields']]}" for d in existing_datatypes])
                    var_str = "\n".join([f"- Variable {v['name']} ({v['logic']})" for v in existing_vars])
                    
                    existing_context_str = f"**EXISTING DATATYPES**:\n{dt_str}\n\n**EXISTING VARIABLES**:\n{var_str}"
                except Exception as e:
                    print(f"[CACHE] Read Error: {e}")

        # 1. Retrieve RAG Context for OpenL Syntax (Functions, Dates, etc.)
        rag_context = self._get_rag_context("OpenL Functions DateUtils BEX Syntax")
        
        parser = JsonOutputParser(pydantic_object=ExtractionResponse)
        # Pass existing_context to prompt
        prompt = PromptTemplate(
            template=ENRICHMENT_PROMPT_TEMPLATE,
            input_variables=["rules", "text", "context", "existing_context"],
            partial_variables={"format_instructions": parser.get_format_instructions()},
        )
        
        chain = prompt | self.llm | parser
        
        # Convert rules to dict if they are objects
        rules_dict = [r.dict() if hasattr(r, 'dict') else r for r in rules]
        
        try:
            # Pass existing_context argument
            result = chain.invoke({
                "rules": rules_dict, 
                "text": text, 
                "context": rag_context,
                "existing_context": existing_context_str
            })
            
            # Post-Processing: Fix common Syntax Hallucinations
            # 1. Fix Dates.diff -> dateDif
            if 'intermediate_variables' in result:
                for v in result['intermediate_variables']:
                    if v.get('logic'):
                        old_logic = v['logic']
                        # Replace Dates.diff(a, b, 'D') with dateDif(a, b, 'D')
                        v['logic'] = v['logic'].replace("Dates.diff", "dateDif")
                        if old_logic != v['logic']:
                            print(f"[DEBUG] Regex Replaced: {old_logic} -> {v['logic']}")
                            
            # 2. Cache Update (Save Result)
            # 2. Cache Update (Merge & Save)
            if filename and cache_path:
                try:
                    # Helper to merge Datatypes
                    def _merge_datatypes(old_list, new_list):
                        dt_map = {d['name']: d for d in old_list}
                        for new_dt in new_list:
                            name = new_dt['name']
                            if name in dt_map:
                                # Merge fields: Add new fields if they don't exist
                                existing_fields = {f['name'] for f in dt_map[name]['fields']}
                                for new_f in new_dt['fields']:
                                    if new_f['name'] not in existing_fields:
                                        dt_map[name]['fields'].append(new_f)
                            else:
                                dt_map[name] = new_dt
                        return list(dt_map.values())

                    # Helper to merge Variables (Overwrite by name)
                    def _merge_variables(old_list, new_list):
                        v_map = {v['name']: v for v in old_list}
                        for new_v in new_list:
                            v_map[new_v['name']] = new_v
                        return list(v_map.values())

                    final_datatypes = result.get("datatypes", [])
                    final_vars = result.get("intermediate_variables", [])
                    
                    if cached_data:
                        final_datatypes = _merge_datatypes(cached_data.get("datatypes", []), final_datatypes)
                        final_vars = _merge_variables(cached_data.get("intermediate_variables", []), final_vars)

                    cache_payload = {
                        "datatypes": final_datatypes,
                        "intermediate_variables": final_vars
                    }
                    
                    with open(cache_path, "w", encoding="utf-8") as f:
                        json.dump(cache_payload, f, indent=2)
                    print(f"[CACHE] Updated and Saved to {cache_path}")
                except Exception as e:
                    print(f"[CACHE] Write Error: {e}")
                        
            return ExtractionResponse(**result)
        except Exception as e:
            print(f"Enrichment failed: {e}")
            raise e

    def _parse_llm_json(self, raw_response: str) -> Any:
        # Helper to clean and parse JSON
        parsed_json = None
        try:
            json_match = re.search(r'\{.*\}|\[.*\]', raw_response, re.DOTALL)
            if json_match:
                raw_response = json_match.group(0)
            parsed_json = json.loads(raw_response)
        except Exception:
            # Simple heuristic repair for trailing commas or markdown
            try:
                raw_response = raw_response.replace("```json", "").replace("```", "").strip()
                parsed_json = json.loads(raw_response)
            except:
                print(f"Failed to parse JSON: {raw_response[:200]}...")
                return {"tables": []}
        
        # Normalize: If it's a list, assume it's a list of tables
        if isinstance(parsed_json, list):
            return {"tables": parsed_json}
        
        # If it's a dict but has no 'tables' key, maybe it IS a table object? 
        # For safety, if it looks like a single table (has 'header' and 'rows'), wrap it.
        if isinstance(parsed_json, dict):
            if "header" in parsed_json and "rows" in parsed_json:
                return {"tables": [parsed_json]}
            
            # Check for generic "tables", "DecisionTables", "Spreadsheets" keys or just use the first list found
            if "tables" in parsed_json:
                return parsed_json
            
            # Fallback: Find first list value
            for key, value in parsed_json.items():
                if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                    return {"tables": value}
                    
            return parsed_json # Return as is, maybe it's valid structure?
            
        return {"tables": []}

    async def generate_excel_structure(self, request: GenerationRequest) -> Dict[str, Any]:
        """
        Execute the 3-Phase Generation Pipeline
        """
        # 1. Phase A: Vocabulary (Datatypes)
        # ----------------------------------
        vocab_context = self._get_rag_context("OpenL Datatype Table syntax and best practices")
        datatypes_input = "\n".join([f"- {d.name}: {[f'{f.name} ({f.type})' for f in d.fields]}" for d in request.datatypes if d.selected])
        print(f"[DEBUG] Datatypes Input to LLM: {datatypes_input}")
        
        prompt_a = PromptTemplate(
            template=DATATYPE_GENERATION_PROMPT_TEMPLATE,
            input_variables=["datatypes_input", "context"]
        )
        chain_a = prompt_a | self.llm
        res_a_raw = chain_a.invoke({"datatypes_input": datatypes_input, "context": vocab_context})
        vocab_structure = self._parse_llm_json(res_a_raw)
        
        # 2. Phase B: Spreadsheets (Calculations)
        # ---------------------------------------
        # We need to identify intermediate variables. 
        # For now, we assume the user might have passed them or we infer them. 
        # But wait, the GenerationRequest currently only has 'rules' and 'datatypes'.
        # The 'IntermediateVariable' model was added to ExtractionResponse but not explicitly to GenerationRequest.
        # However, we can re-derive them or ask the LLM to generate them based on Rule Complexity.
        # Let's ask the LLM to generate Spreadsheets for any complex logic found in the rules.
        
        # 2. Phase B: Spreadsheets (Calculations)
        # ---------------------------------------
        # Use intermediate_variables if provided, otherwise infer from rules (legacy/fallback)
        # [DISABLED] Phase B: Spreadsheets (Calculations)
        # calc_context = self._get_rag_context("OpenL Spreadsheet Table syntax and formulas")
        if request.intermediate_variables:
             variables_text = "\n".join([f"- {v.name} ({v.type}): {v.logic or ''}" for v in request.intermediate_variables])
        else:
             variables_text = "\n".join([f"- {r.name}: {r.condition}" for r in request.rules if r.selected])
        # prompt_b = PromptTemplate(template=SPREADSHEET_GENERATION_PROMPT_TEMPLATE, input_variables=["variables", "context"])
        # chain_b = prompt_b | self.llm
        # res_b_raw = chain_b.invoke({"variables": variables_text, "context": calc_context}) 
        # res_b_raw = res_b_raw.replace("Dates.diff", "dateDif")
        # spreadsheet_structure = self._parse_llm_json(res_b_raw)
        spreadsheet_structure = {} # Phase B Disabled

        # 3. Phase C: Decision Tables (Rules)
        # -----------------------------------
        decision_context = self._get_rag_context("OpenL Decision Table, SmartRules, and Rule Table syntax")
        
        # Filter rules: If rule_type is Spreadsheet/Intermediate, exclude from Phase C?
        # Ideally, Phase C handles "DecisionTable" and "SmartRules".
        # If type is unspecified, include it.
        # If request.intermediate_variables IS provided, then we assume request.rules contains ONLY the Decision Layer rules?
        # Ideally yes. But safekeeping:
        
        decision_rules = [r for r in request.rules if r.selected and r.rule_type in [None, "DecisionTable", "SmartRules", "Rule"]]
        if not decision_rules and not request.intermediate_variables:
             # Fallback: Use all rules if no distinction
             decision_rules = [r for r in request.rules if r.selected]


        
        rules_text_c = "\n".join([f"- ID: {r.id if hasattr(r, 'id') else 'Rule'} | Name: {r.name}: {r.condition}" for r in decision_rules])
        
        prompt_c = PromptTemplate(
            template=DECISION_TABLE_GENERATION_PROMPT_TEMPLATE,
            input_variables=["rules", "datatypes_summary", "variables_summary", "context"]
        )
        chain_c = prompt_c | self.llm
        res_c_raw = chain_c.invoke({
            "rules": rules_text_c, 
            "datatypes_summary": datatypes_input, 
            "variables_summary": variables_text, 
            "context": decision_context
        })
        print(f"[DEBUG] Variables Summary passed to LLM:\n{variables_text}")
        
        # FIX: Enforce dateDif syntax
        res_c_raw = res_c_raw.replace("Dates.diff", "dateDif")
        rules_structure = self._parse_llm_json(res_c_raw)

        # 4. Phase D: Test Generation
        # ---------------------------
        test_context = self._get_rag_context("OpenL Test Table Syntax validation _res_ _error_")
        prompt_d = PromptTemplate(
            template=TEST_GENERATION_PROMPT_TEMPLATE,
            input_variables=["rules_structure", "datatypes_summary", "context"]
        )
        chain_d = prompt_d | self.llm
        
        # Serialize rules structure for context
        rules_structure_str = json.dumps(rules_structure, indent=2)
        
        res_d_raw = chain_d.invoke({
            "rules_structure": rules_structure_str,
            "datatypes_summary": datatypes_input,
            "context": test_context
        })
        test_structure = self._parse_llm_json(res_d_raw)
        tests = test_structure.get("tables", []) if test_structure else []

        # 5. Orchestration / Assembly
        # ---------------------------
        
        spreadsheets = spreadsheet_structure.get("tables", []) if spreadsheet_structure else []
        rules = rules_structure.get("tables", []) if rules_structure else []
        
        # Helper to extract name from header (e.g. "Spreadsheet CalculateRisk" -> "CalculateRisk")
        def get_table_name(table):
            header = table.get("header")
            if not header:
                # Fallback to 'name' key if header is missing (LLM structured output)
                return table.get("name", "")
            
            if isinstance(header, list): 
                header = " ".join(str(h) for h in header)
            
            header_str = str(header)
            parts = header_str.split(" ")
            if len(parts) > 1:
                return parts[-1]
            return header_str

        rule_names = {get_table_name(t) for t in rules}
        # Filter out spreadsheets that have a colliding Decision Table in rules
        unique_spreadsheets = [
            s for s in spreadsheets 
            if get_table_name(s) not in rule_names
        ]

        final_structure = {
            "sheets": [
                {
                    "name": "Vocabulary",
                    "tables": vocab_structure.get("tables", []) if vocab_structure else []
                },
                {
                    "name": "Rules",
                    "tables": unique_spreadsheets + rules
                },
                {
                    "name": "Tests",
                    "tables": tests
                }
            ]
        }
        
        return final_structure

    def create_workbook(self, structure: Dict[str, Any]) -> Workbook:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        sheets = structure.get("sheets", [])
        for sheet_data in sheets:
            ws = wb.create_sheet(sheet_data["name"])
            current_row = 1
            
            for table in sheet_data.get("tables", []):
                # Ensure table is a dictionary
                if not isinstance(table, dict): continue
                
                header_raw = table.get("header") or table.get("Header")
                
                # Construct header if missing
                if header_raw:
                    if isinstance(header_raw, list):
                        header = " ".join(str(h) for h in header_raw)
                    else:
                        header = str(header_raw)
                elif "type" in table and isinstance(table["type"], str) and ("Rules " in table["type"] or "Spreadsheet " in table["type"]):
                    header = table["type"]
                elif "Spreadsheet" in table:
                     header = f"Spreadsheet {table['Spreadsheet']}" if not table['Spreadsheet'].startswith("Spreadsheet") else table['Spreadsheet']
                else:
                    name = table.get("name") or table.get("Name", "")
                    ret_type = table.get("returnType", table.get("type", "Object"))
                    params = table.get("params") or table.get("Params", "")
                    
                    rows = table.get("rows") or table.get("Rows", [])
                    if rows and len(rows) > 0:
                         first_row = rows[0]
                         if isinstance(first_row, list) and len(first_row) > 0 and "Step" in str(first_row[0]):
                             table_type = "Spreadsheet"
                         else:
                             table_type = "Rules"
                    else:
                        table_type = "Rules"
                        
                    header = f"{table_type} {ret_type} {name}({params})"
                
                rows = table.get("rows") or table.get("Rows", [])
                

                
                if not header and not rows: continue
                
                # Write Header
                width = max([len(r) for r in rows]) if rows else 1
                ws.cell(row=current_row, column=1, value=header)
                if width > 1:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=width)
                current_row += 1
                
                # Write Rows
                for row in rows:
                    for col_idx, val in enumerate(row):
                        cell = ws.cell(row=current_row, column=col_idx+1, value=val)
                        # OpenL formulas starting with '=' cause Excel errors if not treated as text
                        # User provided fix: Force String type + quotePrefix
                        if isinstance(val, str) and val.startswith("="):
                            cell.data_type = TYPE_STRING
                            if hasattr(cell, "quotePrefix"):
                                cell.quotePrefix = True
                            elif hasattr(cell, "quote_prefix"):
                                cell.quote_prefix = True
                    current_row += 1
                
                # Spacing
                current_row += 2
                
        return wb
