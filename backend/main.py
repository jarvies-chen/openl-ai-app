from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import shutil
import os
import io
import uuid
from langchain_ollama import OllamaLLM, OllamaEmbeddings
from langchain_core.prompts import PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from dotenv import load_dotenv
from langchain_postgres import PGVector
from langchain_community.document_loaders import PyPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter

load_dotenv()

app = FastAPI(title="OpenL AI App Backend")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
from models import Rule, Datatype, DatatypeField, ExtractionResponse, GenerationRequest, ExtractionRequest, CandidateRule, EnrichmentRequest, CandidateList, SaveVersionRequest, KrakenRuleRequest, KrakenRuleResponse, KrakenDownloadRequest

# Initialize LLM
ollama_base_url = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
ollama_api_key = os.getenv("OLLAMA_API_KEY", "")
llm_model = os.getenv("LLM_MODEL", "gpt-oss:20b")

llm = OllamaLLM(
    base_url=ollama_base_url,
    model=llm_model,
    # headers={"Authorization": f"Bearer {ollama_api_key}"} if ollama_api_key else None
)

@app.get("/")
def read_root():
    return {"message": "OpenL AI App Backend is running"}

from utils import parse_document
from version_control import DocumentManager, RuleDiffer, VersionMetadata, DiffResult

# Initialize DocumentManager
doc_manager = DocumentManager()

@app.post("/upload")
async def upload_document(file: UploadFile = File(...)):
    file_location = f"temp_{file.filename}"
    with open(file_location, "wb+") as file_object:
        shutil.copyfileobj(file.file, file_object)
    
    try:
        text = parse_document(file_location)
        return {"filename": file.filename, "temp_path": file_location, "extracted_text": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/kraken-upload")
async def kraken_upload(file: UploadFile = File(...)):
    file_location = f"temp_kraken_{file.filename}"
    with open(file_location, "wb+") as file_object:
        shutil.copyfileobj(file.file, file_object)
    
    try:
        # Check if it's an Excel file
        _, ext = os.path.splitext(file.filename)
        ext = ext.lower()
        
        if ext not in ['.xlsx', '.xls']:
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
        
        # Parse Excel file
        from utils import parse_excel
        excel_data = parse_excel(file_location)
        
        # Convert to candidate rules format
        candidate_rules = []
        for i, item in enumerate(excel_data["excel_data"], 1):
            candidate_rules.append({
                "id": f"Rule-{i:02d}",
                "name": f"Rule-{i:02d}",
                "summary": item["summary"],
                "source_text": item["source_text"]
            })
        
        # Return in the same format as extract-candidates
        return {
            "filename": file.filename,
            "temp_path": file_location,
            "candidates": candidate_rules
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    # Note: We don't delete temp file here immediately if we want to use it for versioning later, 
    # but for now the flow is: Upload -> Extract -> (User Reviews) -> (User Saves Version?)
    # Actually, the user request implies "Upload -> Save Rules -> Next Upload -> Compare".
    # So we need a way to "Save" the extracted rules as a version.

@app.post("/save-version", response_model=VersionMetadata)
async def save_version(filename: str, request: SaveVersionRequest, temp_path: Optional[str] = None):
    """
    Saves the current set of rules as a new version for the given filename.
    If temp_path is provided, it uses that file. 
    If temp_path is missing but request.text_content is provided (Manual Entry), it creates a temp file from text.
    """
    
    # Handle Manual Entry case (No temp_path, but has text)
    if (not temp_path or not os.path.exists(temp_path)) and request.text_content:
        # Create a temp file from text
        temp_path = f"temp_manual_{uuid.uuid4()}.txt"
        with open(temp_path, "w", encoding="utf-8") as f:
            f.write(request.text_content)
    
    if not temp_path or not os.path.exists(temp_path):
        raise HTTPException(status_code=400, detail="Temporary file not found and no text content provided.")
        
    try:
        version_meta = doc_manager.add_document(temp_path, filename, request.rules, request.comments)
        # Do NOT delete temp file here, as we might need it for subsequent saves (e.g. multiple steps)
        # if os.path.exists(temp_path):
        #     os.remove(temp_path)
        return version_meta
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/documents", response_model=List[dict])
async def get_documents():
    docs = doc_manager.get_documents()
    # Convert to simple list for frontend
    return [
        {
            "base_filename": d.base_filename,
            "version_count": len(d.versions),
            "latest_version": d.versions[-1] if d.versions else None
        }
        for d in docs
    ]

@app.get("/versions/{filename}", response_model=List[VersionMetadata])
async def get_versions(filename: str):
    return doc_manager.get_versions(filename)

@app.get("/rules/{filename}/{version}", response_model=List[Rule])
async def get_rules_version(filename: str, version: int):
    rules = doc_manager.get_rules(filename, version)
    if not rules:
        raise HTTPException(status_code=404, detail="Rules not found for this version")
    return rules

@app.get("/diff/{filename}/{v_old}/{v_new}", response_model=DiffResult)
async def get_diff(filename: str, v_old: int, v_new: int):
    old_rules = doc_manager.get_rules(filename, v_old)
    new_rules = doc_manager.get_rules(filename, v_new)
    
    if not old_rules and v_old != 0: # v_old=0 could mean "nothing"
        raise HTTPException(status_code=404, detail="Old version rules not found")
    if not new_rules:
        raise HTTPException(status_code=404, detail="New version rules not found")
        
    return RuleDiffer.diff(old_rules, new_rules)

@app.get("/rule-history/{filename}/{rule_id}")
async def get_rule_history(filename: str, rule_id: str):
    try:
        history = doc_manager.get_rule_history(filename, rule_id)
        return history
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/document-content/{filename}/{version}")
async def get_document_content(filename: str, version: int):
    try:
        versions = doc_manager.get_versions(filename)
        target_version = next((v for v in versions if v.version == version), None)
        
        if not target_version:
            raise HTTPException(status_code=404, detail="Version not found")
            
        file_path = os.path.join(doc_manager.files_dir, target_version.filename)
        if not os.path.exists(file_path):
             raise HTTPException(status_code=404, detail="File not found")
             
        content = parse_document(file_path)
        return {"content": content}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/extract-rules", response_model=ExtractionResponse)
async def extract_rules(request: ExtractionRequest):
    # Prompt for rule extraction
    parser = JsonOutputParser(pydantic_object=ExtractionResponse)
    prompt = PromptTemplate(
        template="""Analyze the following insurance policy text and extract:
        1. **Business Rules**: Logic statements. 
           - `name`: A meaningful PascalCase name for the rule (e.g., "DetermineEligibility", "CheckFrequencyLimit").
           - `summary`: A brief English description of the rule.
           - `condition`: A simplified pseudo-code representation. For Decision Tables, list the conditions clearly (e.g., "Student=Yes AND Age<=MaxAge").
           - **CRITICAL**: Decompose complex conditions into atomic fields.
             - **WRONG**: A single boolean `isEligibleForSpecialProgram` that hides all the logic.
             - **CORRECT**: Separate fields like `age`, `income`, `enrollmentStatus` so the rule can check them individually.
           - **CRITICAL**: Infer `Integer` for numeric concepts (Age, Duration, Count, Days, Years).
             - **WRONG**: `String employmentDuration`
             - **CORRECT**: `Integer employmentDuration`
           - **CRITICAL**: Group fields into logical **Datatypes** (Entities).
             - **DO NOT** create a separate Datatype for every single field.
             - **WRONG**: `Datatype Age { age }`, `Datatype Income { income }`
             - **CORRECT**: `Datatype Person { age, income, gender }`
             - **Common Groups**:
               - `Policy`: effectiveDate, expirationDate, type, status
               - `Member`: age, gender, employmentStatus, salary
               - `Claim`: amount, date, diagnosisCode
               - `Vehicle`: make, model, year, vin
             - **Action**: Look at the context. If a field belongs to the policy, put it in `Policy`. If it belongs to the person/employee, put it in `Employee`.

        Text:
        {text}

        {format_instructions}
        """,
        input_variables=["text"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )
    
    chain = prompt | llm | parser
    
    try:
        result = chain.invoke({"text": request.text})
        # Ensure all lists are present
        if 'rules' not in result: result['rules'] = []
        if 'datatypes' not in result: result['datatypes'] = []
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/extract-candidates", response_model=List[CandidateRule])
async def extract_candidates(request: ExtractionRequest):
    parser = JsonOutputParser(pydantic_object=CandidateList)
    prompt = PromptTemplate(
        template="""Act as an Insurance Claims Adjuster. Analyze the policy text and extract all business rules relevant to eligibility, coverage, exclusions, and limitations.

        **Goal**: Create a list of "Candidate Rules" in plain English. Do NOT worry about technical syntax or datatypes yet.

        **Instructions**:
        1. Identify every condition that affects a claim's outcome.
        2. Give each rule a meaningful Name (PascalCase).
        3. Write a clear "Summary" in plain English.
        4. Include the "Source Text" snippet.
        5. Assign a unique ID (UUID).

        Text:
        {text}

        {format_instructions}
        """,
        input_variables=["text"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )
    chain = prompt | llm | parser
    try:
        result = chain.invoke({"text": request.text})
        # Result should be {'rules': [...]}
        rules = result.get('rules', [])
        
        # Assign sequential IDs (Rule-01, Rule-02, etc.)
        for i, r in enumerate(rules, 1):
            r['id'] = f"Rule-{i:02d}"
            
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/enrich-rules", response_model=ExtractionResponse)
async def enrich_rules(request: EnrichmentRequest):
    parser = JsonOutputParser(pydantic_object=ExtractionResponse)
    prompt = PromptTemplate(
        template="""Act as an OpenL Tablets Expert. You are given a list of "Candidate Rules" (plain English) and the original policy text.
        
        **Goal**: Convert these candidate rules into fully structured OpenL Rules with technical "Conditions" and "Datatypes".

        **Input Rules**:
        {rules}

        **Context Text**:
        {text}

        **Instructions**:
        1. For each Candidate Rule:
           - Keep the Name and Summary.
           - **Generate the `condition`**: A simplified pseudo-code representation.
             - **CRITICAL**: Decompose complex conditions into atomic fields.
             - **CRITICAL**: Use `&&` and `||` for logic.
             - **CRITICAL**: Use double quotes `"` for strings.
           - **Determine `rule_type`**: 'SmartRules' (field-to-constant) or 'DecisionTable' (variable-to-variable).
        2. **Generate Datatypes**:
           - **CRITICAL**: Group fields into logical Entities (e.g., `Policy`, `Member`, `Claim`).
           - **CRITICAL**: Infer `Integer` for numeric concepts.
           - Ensure ALL fields used in conditions are defined in Datatypes.

        {format_instructions}
        """,
        input_variables=["rules", "text"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )
    chain = prompt | llm | parser
    try:
        # Convert rules to dict for prompt
        rules_dict = [r.dict() for r in request.rules]
        result = chain.invoke({"rules": rules_dict, "text": request.text})
        
        # Ensure lists
        if 'rules' not in result: result['rules'] = []
        if 'datatypes' not in result: result['datatypes'] = []
        
        # Map back IDs from input rules to output rules
        # We assume the LLM maintains the order of rules.
        # If counts mismatch, we try our best or just leave as is.
        input_rules = request.rules
        output_rules = result.get('rules', [])
        
        if len(input_rules) == len(output_rules):
            for in_r, out_r in zip(input_rules, output_rules):
                out_r['id'] = in_r.id
        else:
            # Fallback: Try to match by Name if possible, or just re-assign sequential IDs
            # But for now, let's just ensure they have IDs
            for i, r in enumerate(output_rules, 1):
                if 'id' not in r:
                    r['id'] = f"Rule-{i:02d}"
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/generate-kraken-rules", response_model=KrakenRuleResponse)
async def generate_kraken_rules(request: KrakenRuleRequest):
    try:
        # Read the Kraken rule prompt file
        with open("kraken_rule_prompt.md", "r", encoding="utf-8") as f:
            prompt_template = f.read()
        
        # Format the Excel data into a string
        excel_data_str = "\n".join([f"{item['summary']}\n{item['source_text']}" for item in request.excel_data])
        
        # Combine the prompt template with the Excel data
        full_prompt = f"{prompt_template}\n\nPlease generate the following Kraken rule based on the Kraken rule above:\n\n{excel_data_str}"
        
        # Print full_prompt to log
        print("\n=== Full Prompt for Kraken Rules Generation ===")
        print(full_prompt)
        print("=== End of Full Prompt ===\n")
        
        # Call Ollama API directly without parsing (we want raw text response)
        chain = PromptTemplate(
            template="{prompt}",
            input_variables=["prompt"]
        ) | llm
        
        result = chain.invoke({"prompt": full_prompt})
        print("The result is:\n")
        print(result)
        
        # Return the generated rules as raw text
        return KrakenRuleResponse(generated_rules=result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/kraken-download")
async def kraken_download(request: KrakenDownloadRequest):
    try:
        # Create the content with namespace as first line
        content = f"Namespace {request.name_space}\n\n{request.generated_rules}"
        
        # Ensure the generated directory exists
        os.makedirs("generated", exist_ok=True)
        
        # Generate the file path with the provided file name
        filename = request.file_name
        file_path = os.path.join("generated", filename)
        
        # Write the content to the file
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        
        # Return the download URL
        return {"status": "success", "download_url": f"/download/{filename}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/delete-document/{filename}")
async def delete_document(filename: str):
    success = doc_manager.delete_document(filename)
    if not success:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": f"Document {filename} deleted successfully"}



from openpyxl import Workbook
from langchain_core.runnables import RunnablePassthrough

@app.post("/generate-excel")
async def generate_excel(request: GenerationRequest):
    selected_rules = [r for r in request.rules if r.selected]
    if not selected_rules:
        return {"message": "No rules selected"}
    
    # 1.5 Pre-process rules to enforce Decision Table for complex conditions
    import re
    for rule in selected_rules:
        # Heuristic: If condition contains comparison between two variables (dot notation on both sides)
        # e.g. "policy.effectiveDate < member.hireDate"
        # Regex: word.word ... operator ... word.word
        # Also check for "Date" keyword or specific date logic if needed
        
        condition = rule.condition or ""
        
        # Check for Variable vs Variable comparison
        # Pattern:  word.word  (space/op)  word.word
        # We look for two occurrences of dot-notation separated by an operator
        if re.search(r'[a-zA-Z0-9_]+\.[a-zA-Z0-9_]+\s*[<>=!]+\s*[a-zA-Z0-9_]+\.[a-zA-Z0-9_]+', condition):
            print(f"Forcing DecisionTable for rule '{rule.name}' due to Variable comparison: {condition}")
            rule.rule_type = 'DecisionTable'
            
        # Check for Date vs Date (often implies variable comparison even if not caught above)
        if "Date" in condition or "date" in condition.lower():
             # If it has an operator and isn't just a simple check
             if re.search(r'[<>=!]', condition):
                 # If right side is NOT a simple number or quoted string, assume it's a variable/expression
                 # This is a bit aggressive, but safer for OpenL
                 right_side = re.split(r'[<>=!]+', condition)[-1].strip()
                 if not re.match(r'^[\d"\']', right_side) and right_side.lower() != 'true' and right_side.lower() != 'false':
                     print(f"Forcing DecisionTable for rule '{rule.name}' due to Date logic: {condition}")
                     rule.rule_type = 'DecisionTable'

    # 1. Retrieve OpenL syntax guides from RAG
    vector_store = get_vector_store()
    # Increase k to get more context and use MMR for diversity
    retriever = vector_store.as_retriever(
        search_type="mmr",
        search_kwargs={"k": 10, "fetch_k": 20}
    )
    
    # We query for general OpenL Table syntax and specific rule types
    context_docs = retriever.invoke("OpenL Tablets syntax for Datatype Table, SmartRules, Decision Table, and Spreadsheet structure")
    context_text = "\n\n".join([doc.page_content for doc in context_docs])
    
    # 2. Generate Excel Structure using LLM
    # We ask the LLM to output a logical structure of Tables
    
    class OpenLTable(BaseModel):
        header: str
        rows: List[List[str]] # List of rows, where each row is a list of cell values

    class ExcelSheet(BaseModel):
        name: str
        tables: List[OpenLTable]

    class ExcelStructure(BaseModel):
        sheets: List[ExcelSheet]

    parser = JsonOutputParser(pydantic_object=ExcelStructure)
    
    # Debug: Print RAG context
    try:
        print(f"RAG Context: {context_text.encode('utf-8', errors='ignore').decode('utf-8')}")
    except:
        print("RAG Context: (Error printing context)")

    prompt = PromptTemplate(
        template="""You are an OpenL Tablets expert. Generate the logical structure for an OpenL Excel file based on the following business rules and datatypes.
        
        Business Rules:
        {rules}

        Datatypes:
        {datatypes}
        
        OpenL Syntax Guide (Context):
        {context}
        
        Output a JSON object describing the Excel sheets and their tables.
        
        **CRITICAL**: You MUST generate BOTH sheets:
        1. A "Vocabulary" sheet with all Datatypes
        2. A "Rules" sheet with all the business rules (SmartRules or Decision Tables)
        
        Do NOT skip the Rules sheet. Every business rule provided must be converted into either a SmartRule or Decision Table.
        
        **CRITICAL**: You MUST respect the 'Type' specified for each rule in the input list.
        - The input list has already been pre-processed to force 'DecisionTable' for complex rules.
        - **TRUST THE INPUT TYPE**. If it says 'DecisionTable', generate a Decision Table.
        
        **CRITICAL**: If a rule is marked as 'SmartRules' but involves comparing two variables (e.g., `employee.hours < employee.maxHours`), you MUST generate a **Decision Table** instead.
          - **Detection**: Check the **VALUE** side of the comparison (the part after the operator).
            - **Is it a Variable?** Does it contain a dot `.` (e.g. `e.normalWorkWeek`, `coverage.effectiveDate`)? -> **DECISION TABLE**.
            - **Is it a Math Expression?** Does it contain `+`, `-`, `*`, `/` with a variable? -> **DECISION TABLE**.
            - **Is it `currentDate`?** -> **DECISION TABLE**.
            - **Is it `DateOfService`?** -> **DECISION TABLE**.
            - **Is it a Constant?** (Raw Number `5`, Quoted String `"Active"`, Boolean `true`) -> **SMARTRULES**.
          - **Action**: If ANY variable logic is found on the value side, you MUST generate a **Decision Table**.
          - **CRITICAL**: You must decide the rule type **BEFORE** generating the JSON object.
            - **NEVER** output a SmartRules header and then switch to a Decision Table header inside the same object.
            - **NEVER** output correction text or comments inside the JSON structure.
          - **Decision Table Structure for this case**:
            - Header: `Rules Boolean RuleName(Employee e, Date currentDate)`
            - Row 1: `C1`, `RET1`
            - Row 2: `e.terminationDate <= currentDate + 4`, `result`  (Put the FULL comparison in the expression)
            - Row 3: `Boolean isWithinLimit`, `Boolean result`
            - Row 4: `true`, `true`
            - Row 5: ``, `false`
          - **CRITICAL**: SmartRules ONLY support comparing a field against a **CONSTANT**.
            - **WRONG**: SmartRule with value `< e.normalWorkWeek` (Variable)
            - **WRONG**: SmartRule with value `>= DateOfService` (Variable) -> This results in empty cells like `>=`!
            - **WRONG**: SmartRule with value `>= coverage.effectiveDate` (Variable)
            - **CORRECT**: Decision Table with condition `e.workSchedule < e.normalWorkWeek`
        - If a rule is marked as 'SmartRules' and uses constants:
          - **SmartRules Structure**:
            - Header: `SmartRules <ReturnType> <RuleName>(<Params>)`
            - Row 1: **Column Headers** (Field Paths ONLY).
              - **CRITICAL**: This must be a **SINGLE ROW** containing **ALL** the field paths and the Result column.
              - **CRITICAL**: Do NOT list fields vertically (one per row).
                - **WRONG**: 
                  ```json
                  [
                    ["e.age"], 
                    ["e.isActive"], 
                    ["Result"]
                  ]
                  ```
                - **CORRECT**: 
                  ```json
                  [
                    ["e.age", "e.isActive", "Result"]
                  ]
                  ```
              - **CRITICAL**: This row must contain ONLY the variable/field path (e.g., `employee.age`, `policy.type`, `claim.amount`).
              - **CRITICAL**: Do NOT include operators or values in this row (e.g., `> 18`, `== "Active"`).
              - **WRONG**: `employee.age > 18`
              - **CORRECT**: `employee.age`
              - **CRITICAL**: The LAST column MUST be `Result`.
            - Row 2+: **Values**.
              - Place the operator and value here (e.g., `> 18`, `"Active"`, `true`).
              - **CRITICAL**: The LAST column MUST contain the return value (e.g., `true`, `false`, `"Eligible"`).
                - **WRONG**: `["> 18", "true"]` (Missing return value)
                - **CORRECT**: `["> 18", "true", "true"]` (Includes return value)
              - **CRITICAL**: ALWAYS add a final "Otherwise" row at the bottom.
                - This row should have EMPTY strings `""` for all conditions (meaning "Any").
                - It should return the default value (e.g., `false` for boolean rules).
                - Example: `["", "", "false"]`
              - **CRITICAL**: For Numeric/Date fields, use ONLY the number/date. STRIP ALL UNITS.
                - **WRONG**: `<= 4 Weeks`, `> 18 Years`, `4 Weeks`
                - **CORRECT**: `<= 4`, `> 18`, `4`
              - **CRITICAL**: Comparison operators (>=, <=, >, <) are ONLY for Numbers and Dates.
                - NEVER use them with Strings that contain units.
              - Row 2+ MUST be the lists of values corresponding to those fields.
        - If a rule is marked as 'DecisionTable', you MUST generate a Decision Table.
        
        Requirements:
           - **Structure**:
             - **Header**: `Datatype <Name>` (e.g., `Datatype Driver`)
             - **Rows**: List of `[Type, FieldName]` pairs.
               - **CRITICAL**: This is a VERTICAL table with 2 columns.
               - **Column 1**: The Type (e.g., `String`, `Integer`, `Date`, `Boolean`).
               - **Column 2**: The Field Name (camelCase).
               - **CRITICAL**: You MUST put the **Type** in the first column and the **Name** in the second column.
                 - **WRONG**: `["age", "Integer"]` (Name first)
                 - **CORRECT**: `["Integer", "age"]` (Type first)
               - **CRITICAL**: Do NOT include a header row inside the table (e.g., do NOT add `["Type", "Name"]` or `["Field", "Type"]`). Start directly with the first field definition.
               - **CRITICAL**: Do NOT put multiple fields in one row. Each field gets its own row.
               - **JSON Structure Example for Datatype**:
                 ```json
                 {{
                   "header": "Datatype Driver",
                   "rows": [
                     ["String", "name"],
                     ["Integer", "age"],
                     ["Date", "licenseDate"],
                     ["Boolean", "isActive"]
                   ]
                 }}
                 ```
               - **BAD Example (DO NOT DO THIS)**:
                 ```json
                 {{
                   "header": "Datatype Driver",
                   "rows": [
                     ["Field", "Type"],  // WRONG: Internal header
                     ["name", "String"], // WRONG: Swapped columns
                     ["age", "Integer"]  // WRONG: Swapped columns
                   ]
                 }}
                 ```
                 - **CRITICAL**: For string literals in expressions, ALWAYS use **DOUBLE QUOTES**.
                   - **CORRECT**: `member.roleName == "Employee"`, `status == "Active"`
                   - **WRONG**: `member.roleName == 'Employee'`, `status == 'Active'`
                   - **WRONG**: `e.employmentStatus == 'Terminated'`
                   - **CORRECT**: `e.employmentStatus == "Terminated"`
                   - This applies to ALL string comparisons in Row 2 expressions and Row 4+ values.
                        - **WRONG**: `p.effectiveDate <= Today()`
                        - **WRONG**: `p.effectiveDate <= Today()`
                        - **WRONG**: `p.effectiveDate <= Now()`
                  - **CRITICAL**: **Consistency Check**:
                    - You MUST ensure that EVERY field used in a rule expression is defined in the corresponding Datatype.
                    - If a rule uses `e.isEligibleForDisabilityBenefits`, you MUST add `Boolean isEligibleForDisabilityBenefits` to the `EmploymentDetails` Datatype.
                    - **Do not leave undefined fields.** If you invent a field name for a rule, you MUST add it to the Vocabulary.
                 - **CRITICAL**: NEVER compare a `Date` field with a `Number`.
                   - **WRONG**: `policy.expirationDate > 0`
                   - **WRONG**: `employee.terminationDate <= 30`
                   - **CORRECT STRATEGY**: If checking a duration (e.g. "within 30 days"), you MUST:
                     1. Define an `Integer` field in the Datatype for the duration (e.g. `daysSinceTermination`).
                     2. Use that `Integer` field in the SmartRule column.
                     3. Compare that `Integer` with the number.
                     - Example: Column `e.daysSinceTermination`, Value `<= 30`.
                 - **CRITICAL**: STRIP UNITS from expressions.
                   - **WRONG**: `employee.timeSinceTermination <= 4 weeks`
                   - **CORRECT**: `employee.timeSinceTermination <= 4` (assuming integer represents weeks)
                   - **WRONG**: `age > 18 years`
                   - **CORRECT**: `age > 18`
                   - Ensure the comparison is always `Variable Operator Number` (e.g., `x > 5`), never `Variable Operator Number Unit`.
                 - **CRITICAL**: For custom helper functions (e.g., `isInCosmeticServices()`), you MUST create a Spreadsheet table to define them.
               - Return: 
                 - **CRITICAL**: This must be the **Parameter Name** defined in Row 3 (e.g., `result`), NOT a value like "Approved".
                 - For **Complex Return** (object): Use constructor (e.g., `= new CalculationStatus(flag, code)`).
             - **Row 3 (Parameter Definitions)**: **CRITICAL**: Define the local variable type and name for EVERY COLUMN, including RET1.
               - This is NOT the function parameter - it's a local variable to hold the result of the expression in Row 2.
               - **CRITICAL**: The content MUST be in the format `Type variableName` (e.g., `Boolean isLate`).
               - **CRITICAL**: Do NOT put a description here. Descriptions are NOT allowed in this table structure.
               - **CRITICAL**: Even for complex expressions (e.g., `isInOrthodonticServices(...)`), Row 3 MUST be a variable definition, NOT a description.
                 - **WRONG**: `CDT code is orthodontic` (This is a description)
                 - **CORRECT**: `Boolean isOrthodontic`
                 - **WRONG**: `Date of Service after waiting period` (This is a description)
                 - **CORRECT**: `Boolean isAfterWait`
               - **CRITICAL**: If the content has more than 2 words, it is likely a description and is WRONG. (e.g., "Date of Service within..." is WRONG).
               - **CRITICAL**: You MUST define a parameter for the RET1 column (e.g., `String result`, `Boolean isValid`).
                 - **WRONG**: `Result` (Missing type), `Approved` (Value not parameter).
                 - **CORRECT**: `String result`.
               - **CRITICAL**: Parameter names MUST start with lowercase (e.g., `isValid`, NOT `IsValid`).
               - **CORRECT**: `Boolean isAfterPaid`, `Boolean isWithinGrace`, `String result`.
               - **WRONG**: `Claim claim`, `Policy policy`, `Boolean IsAfterPaid`, `Date of Service after late entrant wait period` (Description is WRONG here), or leaving RET1 empty.
             - **Row 4+ (Values)**: The specific values to check or return.
               - **CRITICAL**: There is NO "Description" row. Row 4 is immediately the first row of values.
               - **CRITICAL**: You MUST generate at least one row of values (Row 4). Do not stop after Row 3.
               - **CRITICAL**: NEVER output a table with only 3 rows. It MUST have at least 4 rows (Header, Expressions, Variables, Values).
               - **CRITICAL**: If the rule logic implies a boolean condition (e.g., "If A and B then True"), you MUST generate the Truth Table rows covering the "True" case.
                 - If specific values are not mentioned, generate the "Happy Path" row where the result is True.
                 - **CRITICAL**: If the rule logic implies a boolean condition (e.g., "If A and B then True"), you MUST generate the Truth Table rows covering the "True" case.
                 - If specific values are not mentioned, generate the "Happy Path" row where the result is True.
                 - Example: `["true", "true", "true"]`
               - **CRITICAL**: ALWAYS add a final "Otherwise" row at the bottom.
                 - This row should have EMPTY strings `""` for all conditions (meaning "Any").
                 - It should return the default value (e.g., `false` for boolean rules).
               - **CRITICAL**: String values must be quoted (e.g., `"Approved"`, `"Denied"`).
           - **Example (Splitting AND Conditions)**:
             - Rule: `Member.recordExists && (DateOfService >= Member.hireDate)`
             - Header: `Rules String CheckEligibility(Member m, Date dos)`
             - Row 1: `["C1", "C2", "RET1"]`
             - Row 2: `["m.recordExists", "dos >= m.hireDate", "result"]`
             - Row 3: `["Boolean exists", "Boolean isEligible", "String result"]`
             - Row 4: `["true", "true", "\\"Approved\\""]`
             - Row 5: `["", "", "\\"Denied\\""]` // Otherwise row
             
           - **CRITICAL**: Operators MUST be Java/C style.
             - **WRONG**: `AND`, `OR`
             - **CORRECT**: `&&`, `||`
           - **CRITICAL**: Split complex conditions into separate columns.
             - **WRONG**: `C1: condition1 && condition2` (Single column)
             - **CORRECT**: `C1: condition1`, `C2: condition2` (Two columns)
        
        3. **Spreadsheet Tables (Helper Functions)**:
           - **CRITICAL**: If you need custom helper functions (e.g., `isInCosmeticServices()`, `contains()`), create Spreadsheet tables to define them.
           - Place Spreadsheet tables in the **Rules** sheet, BEFORE the Decision Tables that use them.
           - **Header**: `Spreadsheet <ReturnType> <FunctionName>(<Params>)`
           - **Structure**:
             - **Row 1**: Column headers (e.g., `Step`, `Name`, `Value`)
             - **Row 2+**: Steps defining the function logic
           - **Example (Helper Function)**:
             - Function: `isInCosmeticServices(String cdtCode)`
             - Header: `Spreadsheet Boolean isInCosmeticServices(String cdtCode)`
             - Row 1: `["Step", "Name", "Value"]`
             - Row 2: `["1", "cosmeticCodes", "Arrays.asList(\\"D9970\\", \\"D9971\\", \\"D9972\\")"]`
             - Row 3: `["2", "result", "cosmeticCodes.contains(cdtCode)"]`
             - Row 4: `["3", "return", "result"]`
        
        {format_instructions}
        """,
        input_variables=["rules", "datatypes", "context"],
        partial_variables={"format_instructions": parser.get_format_instructions()},
    )
    
    # chain = prompt | llm | parser
    chain_raw = prompt | llm
    
    rules_text = "\n".join([f"- Name: {r.name}, Summary: {r.summary}, Type: {r.rule_type} (Condition: {r.condition}, Result: {r.result})" for r in selected_rules])
    datatypes_text = "\n".join([f"- {d.name}: {[f.name for f in d.fields]}" for d in request.datatypes if d.selected])
    
    try:
        print(f"Generating Excel with rules: {rules_text}")
        # Get raw string response first
        raw_response = chain_raw.invoke({
            "rules": rules_text, 
            "datatypes": datatypes_text,
            "context": context_text
        })
        
        # Print first 500 chars for debugging
        print(f"Raw LLM Response (first 500 chars): {raw_response[:500]}")
        print(f"Raw LLM Response (last 100 chars): {raw_response[-100:]}")
        
        if not raw_response:
             raise ValueError("LLM returned empty string.")

        import re
        import json
        # Clean the response to ensure it's valid JSON
        json_match = re.search(r'\{.*\}', raw_response, re.DOTALL)
        if json_match:
            raw_response = json_match.group(0)
        
        # Parse the raw response
        structure_data = None
        parse_error = None
        
        # Try 1: Use the Pydantic parser
        try:
            structure_data = parser.parse(raw_response)
        except Exception as e:
            parse_error = e
            print(f"Pydantic Parse Error: {e}")
        
        # Try 2: Use json.loads as fallback
        if not structure_data:
            try:
                json_data = json.loads(raw_response)
                structure_data = json_data
                print("Successfully parsed with json.loads")
            except Exception as e:
                print(f"json.loads Parse Error: {e}")
        
        # Try 3: Heuristic repair
        if not structure_data:
            print("Parser returned None or failed, attempting heuristic repair...")
            cleaned_response = raw_response.strip()
            success = False
            
            # Try removing trailing characters one by one
            for i in range(5):
                if not cleaned_response: break
                cleaned_response = cleaned_response[:-1]
                
                try:
                    json_data = json.loads(cleaned_response)
                    structure_data = json_data
                    success = True
                    print(f"Successfully parsed JSON after removing {i+1} trailing char(s)")
                    break
                except:
                    continue
            
            if not success:
                print(f"FULL Raw Response for debugging:\n{raw_response}")
                raise ValueError(f"Failed to parse JSON after all retries. Original error: {parse_error}")
        
        print(f"LLM Response Structure: {structure_data}")
        
        if not structure_data:
            raise ValueError("Failed to generate Excel structure: LLM returned empty response.")

        # 3. Create Excel File
        wb = Workbook()
        # Keep default "Sheet" for now, remove it later if we add new sheets
        
        # Normalize the structure to handle case variations and different formats
        def normalize_dict(d):
            """Convert all keys to lowercase"""
            if isinstance(d, dict):
                return {k.lower(): normalize_dict(v) for k, v in d.items()}
            elif isinstance(d, list):
                return [normalize_dict(item) for item in d]
            else:
                return d
        
        structure_data = normalize_dict(structure_data)
            
        # Handle case where structure_data might be a list (if LLM ignored the root object)
        sheets_data = structure_data.get('sheets', []) if isinstance(structure_data, dict) else []
        
        # Handle alternative format where LLM returns {Vocabulary: {...}, Rules: {...}}
        if not sheets_data and isinstance(structure_data, dict):
            # Check if the structure has 'vocabulary' and 'rules' as top-level keys
            if 'vocabulary' in structure_data or 'rules' in structure_data:
                sheets_data = []
                if 'vocabulary' in structure_data:
                    vocab_data = structure_data['vocabulary']
                    sheets_data.append({
                        'name': vocab_data.get('name', 'Vocabulary'),
                        'tables': vocab_data.get('tables', [])
                    })
                if 'rules' in structure_data:
                    rules_data = structure_data['rules']
                    sheets_data.append({
                        'name': rules_data.get('name', 'Rules'),
                        'tables': rules_data.get('tables', [])
                    })
        
        if not sheets_data and isinstance(structure_data, list):
             sheets_data = structure_data

        for sheet_data in sheets_data:
            # Parse sheet data
            name = sheet_data.get('name', 'Sheet') if isinstance(sheet_data, dict) else getattr(sheet_data, 'name', 'Sheet')
            tables = sheet_data.get('tables', []) if isinstance(sheet_data, dict) else getattr(sheet_data, 'tables', [])
            
            ws = wb.create_sheet(name)
            current_row = 1
            
            for table in tables:
                # Parse table data
                header = table.get('header', '') if isinstance(table, dict) else getattr(table, 'header', '')
                rows = table.get('rows', []) if isinstance(table, dict) else getattr(table, 'rows', [])
                
                if not header and not rows:
                    continue

                # Calculate width for merging header
                # Width is max of row length, or 1 if empty
                width = 1
                if rows:
                    width = max([len(r) for r in rows]) if rows else 1
                
                # Write Header
                c = ws.cell(row=current_row, column=1, value=header)
                if width > 1:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=width)
                
                # Apply header style (optional, but good for OpenL)
                # c.font = Font(bold=True, color="FFFFFF")
                # c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

                current_row += 1
                
                # Write Rows
                for row_data in rows:
                    for col_idx, cell_value in enumerate(row_data):
                        ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                    current_row += 1
                
                # Add spacing between tables (2 empty rows)
                current_row += 2
        
        # Remove default "Sheet" if we created other sheets
        if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Save to temp file
        filename = f"OpenL_Rules_{uuid.uuid4()}.xlsx"
        save_path = os.path.join("generated", filename)
        os.makedirs("generated", exist_ok=True)
        wb.save(save_path)
        
        # Return the relative download URL
        # Frontend will prepend the base URL
        return {"status": "success", "download_url": f"/download/{filename}"}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error generating Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{filename}")
async def download_file(filename: str):
    file_path = os.path.join("generated", filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename)
    raise HTTPException(status_code=404, detail="File not found")
        
embedding_model = os.getenv("EMBEDDING_MODEL", "mxbai-embed-large:latest")

embeddings = OllamaEmbeddings(
    base_url=ollama_base_url,
    model=embedding_model,
    # headers={"Authorization": f"Bearer {ollama_api_key}"} if ollama_api_key else None
)

# Connection string for pgvector
DB_USER = os.getenv("PG_USER", "user")
DB_PASSWORD = os.getenv("PG_PASSWORD", "password")
DB_HOST = os.getenv("PG_HOST", "localhost")
DB_PORT = os.getenv("PG_PORT", "5432")
DB_NAME = os.getenv("PG_DB", "openl_rag")

CONNECTION_STRING = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

def get_vector_store():
    return PGVector(
        embeddings=embeddings,
        collection_name="openl_guide",
        connection=CONNECTION_STRING,
        use_jsonb=True,
    )

@app.post("/ingest-guide")
async def ingest_guide(file: UploadFile = File(...)):
    file_location = f"temp_guide_{file.filename}"
    with open(file_location, "wb+") as file_object:
        shutil.copyfileobj(file.file, file_object)
    
    try:
        loader = PyPDFLoader(file_location)
        docs = loader.load()
        
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
        splits = text_splitter.split_documents(docs)
        
        vector_store = get_vector_store()
        vector_store.add_documents(splits)
        
        return {"message": "Ingestion complete", "chunks": len(splits)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(file_location):
            os.remove(file_location)
