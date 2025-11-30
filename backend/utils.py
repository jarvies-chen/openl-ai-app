import os
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

def parse_pdf(file_path: str) -> str:
    try:
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error parsing PDF: {e}")
        return ""

def parse_docx(file_path: str) -> str:
    try:
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text
    except Exception as e:
        print(f"Error parsing DOCX: {e}")
        return ""

def parse_excel(file_path: str) -> dict:
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Read C and D columns
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Start from row 2 (skip header)
            c_col = row[2] if len(row) > 2 else ""
            d_col = row[3] if len(row) > 3 else ""
            
            if c_col or d_col:  # Only add rows with data
                data.append({
                    "summary": str(c_col) if c_col else "",
                    "source_text": str(d_col) if d_col else ""
                })
        
        return {"excel_data": data}
    except Exception as e:
        print(f"Error parsing Excel: {e}")
        return {"excel_data": []}

def parse_document(file_path: str) -> str:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    if ext == '.pdf':
        return parse_pdf(file_path)
    elif ext == '.docx':
        return parse_docx(file_path)
    elif ext == '.txt':
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print(f"Error parsing TXT: {e}")
            return ""
    else:
        raise ValueError(f"Unsupported file format: {ext}")
